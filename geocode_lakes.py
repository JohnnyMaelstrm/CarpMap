"""
geocode_lakes.py
----------------
Aja kerran: python3 geocode_lakes.py
Tallentaa koordinaatit lakes_geocoded.json tiedostoon.
Kestää ~2 min (Nominatim rate limit), mutta tarvitsee ajaa vain kerran
tai kun lisäät uusia Excel-tiedostoja.
"""

import openpyxl, json, time, urllib.request, urllib.parse, re

# ---- MUOKKAA TÄTÄ LISTAA KUN TULEE UUSIA EXCELEITÄ ----
SOURCES = [
    ('Istutukset_Karppi-Siika_Pirkanmaa.xlsx',       'Pirkanmaa'),
    ('Istutukset_Siika-Karppi_Varsinais-Suomi.xlsx',  'Varsinais-Suomi'),
    # ('Uusi_alue.xlsx', 'Uusi Alue'),
]

def extract_carp(filepath, region):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    lakes = {}
    for row in ws.iter_rows(min_row=10, max_row=ws.max_row, values_only=True):
        laji = row[15] if len(row) > 15 else None
        if not (laji and 'arppi' in str(laji).lower()): # Varmistetaan että löytyy iso tai pieni K
            continue
            
        jarvinumero = str(row[6]).strip() if row[6] else ''
        nimi = str(row[5]).strip() if row[5] else ''
        kunta = str(row[10]).strip() if row[10] else ''
        
        # Luodaan yksilöllinen avain (järvinumero on paras, jos puuttuu niin nimi+kunta)
        key = jarvinumero if jarvinumero and jarvinumero != 'None' else f"{nimi}_{kunta}"
        key = key + '_' + region
        
        if key not in lakes:
            lakes[key] = {
                'jarvinumero': jarvinumero,
                'nimi': nimi,
                'kunta': kunta,
                'region': region,
                'lat': None,
                'lng': None,
                'istutukset': []
            }
        kpl = row[28]
        lakes[key]['istutukset'].append({
            'vuosi': row[1],
            'pvm': str(row[4])[:10] if row[4] else '',
            'kpl': int(kpl) if kpl else 0,
            'ika': str(row[17]) if row[17] else '',
            'paino_g': float(row[21]) if row[21] else None,
        })
    return list(lakes.values())

def clean_name(nimi):
    return re.sub(r'\s*\(N60.*?\)x?\d*', '', nimi).replace('(kesk. N60+78.10)', '').strip()

def hae_jarviwiki(jarvinumero):
    """Hakee absoluuttisen tarkan koordinaatin SYKEn Järviwiki-rajapinnasta"""
    if not jarvinumero or jarvinumero == 'None' or str(jarvinumero).strip() == '':
        return None, None
    url = f"https://www.jarviwiki.fi/api.php?action=ask&query=[[J%C3%A4rvinumero::{urllib.parse.quote(str(jarvinumero))}]]|?Koordinaatti_lat|?Koordinaatti_pit&format=json"
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'CarpMap/1.0'})
        with urllib.request.urlopen(req) as response:
            data = json.loads(response.read())
            results = data.get('query', {}).get('results', {})
            if results:
                first_key = list(results.keys())[0]
                printouts = results[first_key].get('printouts', {})
                lat = printouts.get('Koordinaatti lat', [None])[0]
                lon = printouts.get('Koordinaatti pit', [None])[0]
                if lat and lon:
                    return float(lat), float(lon)
    except Exception as e:
        pass
    return None, None

def geocode_nominatim(nimi, kunta, region):
    """Varajärjestelmä (Merialueet ja nimettömät lammet)"""
    name = clean_name(nimi)
    kunta_cap = kunta.capitalize()
    
    # Optimoitu hakujärjestys
    queries = [
        f"{name} {kunta_cap} Finland",
        f"{name} {region} Finland",
        f"{name} Finland",
    ]
    for q in queries:
        url = 'https://nominatim.openstreetmap.org/search?' + urllib.parse.urlencode({
            'q': q, 'format': 'json', 'limit': 3, 'countrycodes': 'fi'
        })
        req = urllib.request.Request(url, headers={
            'User-Agent': 'CarpMap/1.0',
            'Accept-Language': 'fi'
        })
        try:
            with urllib.request.urlopen(req, timeout=10) as resp:
                data = json.loads(resp.read())
            if data:
                water = next((d for d in data if d.get('type') in ('water','lake','reservoir', 'sea') or d.get('class') in ('natural','water')), None)
                hit = water or data[0]
                return float(hit['lat']), float(hit['lon'])
        except Exception as e:
            pass # Hiljennä virheet, kokeillaan seuraavaa
        time.sleep(1.1)
    return None, None

# ---- MAIN ----
all_lakes_dict = {} # Käytetään sanakirjaa päällekkäisyyksien poistamiseen
for filepath, region in SOURCES:
    lakes_from_file = extract_carp(filepath, region)
    for lake in lakes_from_file:
        key = lake['jarvinumero'] + '_' + lake['region'] + '_' + lake['kunta']
        if key in all_lakes_dict:
             all_lakes_dict[key]['istutukset'].extend(lake['istutukset'])
        else:
             all_lakes_dict[key] = lake

all_lakes = list(all_lakes_dict.values())
print(f"Found {len(all_lakes)} unique lakes total. Starting geocoding...\n")

# Load existing results if available (resume support)
try:
    with open('lakes_geocoded.json', 'r', encoding='utf-8') as f:
        # Käytetään samaa avainlogiikkaa täälläkin
        existing = {(l['jarvinumero'] + '_' + l['region'] + '_' + l['kunta']): l for l in json.load(f)}
    print(f"Resuming — {sum(1 for l in existing.values() if l['lat'])} already geocoded\n")
except FileNotFoundError:
    existing = {}

ok = 0
failed = []

for i, lake in enumerate(all_lakes):
    key = lake['jarvinumero'] + '_' + lake['region'] + '_' + lake['kunta']
    
    # Reuse existing coordinate if already found
    if key in existing and existing[key]['lat']:
        lake['lat'] = existing[key]['lat']
        lake['lng'] = existing[key]['lng']
        ok += 1
        continue

    print(f"[{i+1}/{len(all_lakes)}] {lake['nimi']} ({lake['kunta']})...", end=' ', flush=True)
    
    # 1. Yritetään täydellistä osumaa Järvinumerolla
    lat, lng = hae_jarviwiki(lake['jarvinumero'])
    lähde = "Järviwiki"
    
    # 2. Jos ei onnistunut, käytetään Nominatimia
    if not lat:
        lat, lng = geocode_nominatim(lake['nimi'], lake['kunta'], lake['region'])
        lähde = "Nominatim"
        
    if lat:
        lake['lat'] = lat
        lake['lng'] = lng
        print(f"✓ {lat:.4f}, {lng:.4f} [{lähde}]")
        ok += 1
    else:
        lake['lat'] = None
        lake['lng'] = None
        print("✗ not found")
        failed.append(lake['nimi'])
    
    time.sleep(1.2) # Järviwikin takia ei haittaa pitää viivettä

# Save results
with open('lakes_geocoded.json', 'w', encoding='utf-8') as f:
    json.dump(all_lakes, f, ensure_ascii=False, indent=2)

print(f"\nDone! {ok}/{len(all_lakes)} geocoded.")
if failed:
    print(f"Not found ({len(failed)}): {', '.join(failed)}")
print("\nNext: run  python3 merge_data.py  to update the HTML.")